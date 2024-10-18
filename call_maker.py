import requests
import json
from openpyxl import load_workbook

# Function to read phone numbers from Excel file
def read_phone_numbers_from_excel(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        data.append({'phoneNumber': row[0], 'Response': row[1]})
    return data, workbook, sheet

# Function to update the Excel sheet with response status
def update_excel_sheet(workbook, sheet, data):
    for idx, row in enumerate(data, start=2):  # Start from the second row
        sheet.cell(row=idx, column=2, value=row['Response'])
    workbook.save('path/to/excel/file.xlsx')

# Function to make outbound calls and process responses
def make_outbound_calls(file_path):
    data, workbook, sheet = read_phone_numbers_from_excel(file_path)

    for row in data:
        caller_number = row['Phone Number']
        try:
            response = requests.post(
                'https://example.com/webhook',
                headers={'Content-Type': 'application/json'},
                data=json.dumps({
                    'route': "1",  # Route 1 is for getting the first message
                    'data1': caller_number,  # Send caller's number, this will actually find about the past conversation with that number.
                    'data2': "empty"  # Extra data (not used here) but we can add about that person like personality, current work for personalization.
                })
            )

            if response.ok:
                response_text = response.text  # Get the text response from the make.com webhook
                print('Make.com webhook response:', response_text)
                try:
                    response_data = response.json()  # Try to parse the response from webhook as JSON
                    if response_data and 'firstMessage' in response_data:
                        first_message = response_data['firstMessage']  # If there's a firstMessage in the response, use it
                        print('Parsed firstMessage from Make.com:', first_message)
                    else:
                        first_message = response_text.strip()  # Use the plain text response if parsing fails
                except json.JSONDecodeError as parse_error:
                    print('Error parsing webhook response:', parse_error)  # Log any errors while parsing the response
                    first_message = response_text.strip()  # Use the plain text response if parsing fails
            else:
                print('Failed to send data to Make.com webhook:', response.status_code, response.reason)  # Log if webhook fails
                row['Response'] = 'No answer'
        except requests.RequestException as error:
            print('Error sending data to Make.com webhook:', error)  # Log if an error occurs in the request
            row['Response'] = 'Error'

        # Set up a new session for this call
        session = {
            'transcript': '',  # Store the conversation transcript here
            'streamSid': None  # This will be set when the media stream starts
        }

    # Update the Excel sheet with response status
    update_excel_sheet(workbook, sheet, data)

# Main function
def main():
    file_path = 'path/to/excel/file.xlsx'
    make_outbound_calls(file_path)

if __name__ == '__main__':
    main()